"""
CSV/Excel Exporter
Export data in CSV and Excel formats
"""

import csv
import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("aiobs.exports.csv")


class CSVExporter:
    """Export data to CSV and Excel formats"""

    def __init__(self):
        # GASKIA brand colors for Excel
        self.header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.accent_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
        self.border = Border(
            left=Side(style='thin', color='E0E0DA'),
            right=Side(style='thin', color='E0E0DA'),
            top=Side(style='thin', color='E0E0DA'),
            bottom=Side(style='thin', color='E0E0DA'),
        )

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
    ) -> bytes:
        """
        Export data to CSV format.

        Args:
            data: List of dictionaries to export
            columns: Optional list of columns to include (in order)

        Returns:
            CSV bytes
        """
        if not data:
            return b""

        buffer = io.StringIO()

        # Determine columns
        if columns is None:
            columns = list(data[0].keys())

        writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()

        for row in data:
            # Convert complex types to strings
            clean_row = {}
            for col in columns:
                value = row.get(col, "")
                if isinstance(value, (list, dict)):
                    value = str(value)
                elif isinstance(value, datetime):
                    value = value.isoformat()
                clean_row[col] = value
            writer.writerow(clean_row)

        return buffer.getvalue().encode('utf-8')

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        sheet_name: str = "Data",
        title: Optional[str] = None,
    ) -> bytes:
        """
        Export data to Excel format with GASKIA styling.

        Args:
            data: List of dictionaries to export
            columns: Optional list of columns to include
            sheet_name: Name of the worksheet
            title: Optional title row

        Returns:
            Excel bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        current_row = 1

        # Add title if provided
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns or list(data[0].keys())))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(size=14, bold=True, color="1A1A2E")
            title_cell.alignment = Alignment(horizontal='center')
            current_row = 3

        if not data:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        # Determine columns
        if columns is None:
            columns = list(data[0].keys())

        # Write header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name.replace('_', ' ').title())
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        current_row += 1

        # Write data rows
        for row_data in data:
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")

                # Convert complex types
                if isinstance(value, (list, dict)):
                    value = str(value)
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')

                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')

            current_row += 1

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row in range(2, min(current_row, 100)):  # Sample first 100 rows
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        # Freeze header row
        ws.freeze_panes = ws.cell(row=2 if not title else 4, column=1)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def export_metrics(
        self,
        model_id: str,
        metrics: List[Dict[str, Any]],
        format: str = "csv",
    ) -> bytes:
        """
        Export model metrics.

        Args:
            model_id: Model identifier
            metrics: List of metric data points
            format: 'csv' or 'xlsx'

        Returns:
            Export bytes
        """
        columns = ['timestamp', 'metric_name', 'value', 'labels']

        if format == "xlsx":
            return self.export_to_excel(
                metrics,
                columns=columns,
                sheet_name="Metrics",
                title=f"Model Metrics - {model_id}",
            )
        else:
            return self.export_to_csv(metrics, columns=columns)

    async def export_audit_trail(
        self,
        entries: List[Dict[str, Any]],
        format: str = "csv",
    ) -> bytes:
        """
        Export audit trail entries.

        Args:
            entries: List of audit entries
            format: 'csv' or 'xlsx'

        Returns:
            Export bytes
        """
        columns = ['timestamp', 'actor', 'action', 'resource', 'outcome', 'context']

        if format == "xlsx":
            return self.export_to_excel(
                entries,
                columns=columns,
                sheet_name="Audit Trail",
                title="AIOBS Audit Trail Export",
            )
        else:
            return self.export_to_csv(entries, columns=columns)

    async def export_compliance_data(
        self,
        data: Dict[str, Any],
        format: str = "csv",
    ) -> bytes:
        """
        Export compliance assessment data.

        Args:
            data: Compliance data
            format: 'csv' or 'xlsx'

        Returns:
            Export bytes
        """
        # Flatten compliance data into rows
        rows = []

        for model in data.get('models', []):
            rows.append({
                'model_id': model.get('id'),
                'model_name': model.get('name'),
                'risk_level': model.get('risk_level'),
                'trust_score': model.get('trust_score'),
                'compliance_status': model.get('compliance_status'),
                'last_assessed': model.get('last_assessed'),
                'findings_count': len(model.get('findings', [])),
            })

        columns = ['model_id', 'model_name', 'risk_level', 'trust_score',
                   'compliance_status', 'last_assessed', 'findings_count']

        if format == "xlsx":
            return self.export_to_excel(
                rows,
                columns=columns,
                sheet_name="Compliance",
                title="AIOBS Compliance Assessment",
            )
        else:
            return self.export_to_csv(rows, columns=columns)

    async def export_alerts(
        self,
        alerts: List[Dict[str, Any]],
        format: str = "csv",
    ) -> bytes:
        """
        Export alert history.

        Args:
            alerts: List of alert data
            format: 'csv' or 'xlsx'

        Returns:
            Export bytes
        """
        columns = ['id', 'title', 'severity', 'type', 'status',
                   'model_id', 'service_id', 'triggered_at', 'resolved_at']

        if format == "xlsx":
            return self.export_to_excel(
                alerts,
                columns=columns,
                sheet_name="Alerts",
                title="AIOBS Alert History",
            )
        else:
            return self.export_to_csv(alerts, columns=columns)
